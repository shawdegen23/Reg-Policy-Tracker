#!/usr/bin/env python3
"""
Regenerate index.html from the TEC regulatory tracker spreadsheet.

Workflow:
  1. Edit ../TEC_Regulatory_Tracker.xlsx (the source of truth)
  2. python3 build_site.py
  3. git add index.html && git commit -m "Update tracker" && git push

The spreadsheet is the single source of truth; this script bakes its data
into a self-contained index.html so the Vercel site always matches the tracker.
"""
import json, os, datetime
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "..", "TEC_Regulatory_Tracker.xlsx")
OUT = os.path.join(HERE, "index.html")

TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>TEC Regulatory Monitoring Dashboard</title>
<style>
  :root{--navy:#1f3b57;--blue:#2e5e8c;--lt:#dce6f1;--bg:#f4f7fa;--card:#fff;--ink:#1c2b38;--muted:#5b6b78;--line:#e2e8ee;--amber:#e0a800;--amberbg:#fff6da;--green:#2f8f4e;--greenbg:#e6f4ea;--red:#c0392b;--redbg:#fbe9e7}
  *{box-sizing:border-box}
  body{margin:0;font-family:-apple-system,Segoe UI,Arial,sans-serif;background:var(--bg);color:var(--ink);line-height:1.45}
  header{background:linear-gradient(135deg,var(--navy),var(--blue));color:#fff;padding:26px 32px}
  header h1{margin:0;font-size:22px}
  header p{margin:6px 0 0;opacity:.85;font-size:13px}
  .wrap{max-width:1200px;margin:0 auto;padding:24px 32px 60px}
  .bar{display:flex;flex-wrap:wrap;gap:8px;margin:18px 0 6px}
  .bar a{font-size:12.5px;text-decoration:none;color:var(--blue);background:#fff;border:1px solid var(--line);padding:7px 12px;border-radius:20px;transition:.15s}
  .bar a:hover{background:var(--blue);color:#fff}
  .tabs{display:flex;gap:4px;margin-top:22px;border-bottom:2px solid var(--line);flex-wrap:wrap}
  .tab{padding:10px 18px;cursor:pointer;font-size:14px;font-weight:600;color:var(--muted);border-bottom:3px solid transparent;margin-bottom:-2px}
  .tab.active{color:var(--navy);border-bottom-color:var(--blue)}
  section{display:none;padding-top:20px}
  section.active{display:block}
  .grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(340px,1fr));gap:16px}
  .card{background:var(--card);border:1px solid var(--line);border-radius:12px;padding:16px 18px;box-shadow:0 1px 2px rgba(0,0,0,.03)}
  .card h3{margin:0 0 4px;font-size:15px}
  .card .docket{font-size:12px;color:var(--muted);font-weight:600}
  .tag{display:inline-block;font-size:11px;font-weight:700;padding:3px 9px;border-radius:12px;margin-top:8px}
  .full{background:var(--amberbg);color:#8a6100;border:1px solid #f0d98a}
  .scan{background:var(--lt);color:var(--navy)}
  .prio{background:var(--greenbg);color:var(--green);border:1px solid #b7ddc2}
  .confirm{background:var(--redbg);color:var(--red);border:1px solid #f0b7ae}
  .cat{font-size:11.5px;color:var(--muted);margin-top:8px}
  .desc{font-size:13px;margin-top:8px}
  .kv{font-size:12px;color:var(--muted);margin-top:8px}
  .kv b{color:var(--ink)}
  .card a.open{display:inline-block;margin-top:12px;font-size:12.5px;font-weight:600;text-decoration:none;color:#fff;background:var(--blue);padding:7px 12px;border-radius:8px}
  .card a.open:hover{background:var(--navy)}
  .sub{font-size:12px;color:var(--muted);margin-top:10px}
  .sub a{color:var(--blue)}
  .note{background:#fff;border-left:4px solid var(--amber);padding:14px 16px;border-radius:8px;margin:8px 0 22px;font-size:13px}
  .note b{color:var(--navy)}
  table{width:100%;border-collapse:collapse;background:#fff;border-radius:10px;overflow:hidden;font-size:13px}
  th{background:var(--navy);color:#fff;text-align:left;padding:11px 13px;font-size:12.5px}
  td{padding:11px 13px;border-top:1px solid var(--line);vertical-align:top}
  tr:nth-child(even) td{background:#f8fafc}
  tr.prio-row td{background:#fff9e8}
  .pill{font-size:11px;font-weight:700;padding:2px 8px;border-radius:10px}
  .pill.High{background:var(--redbg);color:var(--red)}
  .pill.Medium{background:var(--amberbg);color:#8a6100}
  .pill.Low,.pill.Monitor{background:var(--lt);color:var(--navy)}
  .search{width:100%;max-width:360px;padding:9px 12px;border:1px solid var(--line);border-radius:8px;font-size:13px;margin-bottom:14px}
  .empty{color:var(--muted);font-size:13px;padding:16px}
  footer{color:var(--muted);font-size:12px;text-align:center;padding:26px}
</style>
</head>
<body>
<header>
  <h1>TEC Regulatory Monitoring Dashboard</h1>
  <p>Regulatory Policy &amp; Strategy · Powered by the master tracker spreadsheet · Updated <span id="upd"></span></p>
</header>
<div class="wrap">
  <div class="bar">
    <a href="https://apps.cpuc.ca.gov/apex/f?p=401:1:0" target="_blank">CPUC Proceedings</a>
    <a href="https://docs.cpuc.ca.gov/advancedsearchform.aspx" target="_blank">CPUC Doc Search</a>
    <a href="https://www.cpuc.ca.gov/proceedings-and-rulemaking" target="_blank">CPUC New Filings</a>
    <a href="https://www.energy.ca.gov/proceedings" target="_blank">CEC</a>
    <a href="https://ww2.arb.ca.gov/rulemaking-activity" target="_blank">CARB</a>
    <a href="https://www.aqmd.gov/nav/rules" target="_blank">SCAQMD</a>
    <a href="https://leginfo.legislature.ca.gov/faces/billSearchClient.xhtml" target="_blank">Leginfo</a>
  </div>
  <div class="tabs">
    <div class="tab active" data-t="proc">Proceedings</div>
    <div class="tab" data-t="digest">Digest Log</div>
    <div class="tab" data-t="agency">Agency Watchlist</div>
    <div class="tab" data-t="subs">Subscriptions</div>
  </div>

  <section id="proc" class="active">
    <div class="note"><b>How to use:</b> Each card links to the live CPUC docket. R.25-04-010 (Energy Efficiency) is the full-read proceeding; the rest are scan-for-relevance. Edit the spreadsheet and rerun <code>build_site.py</code> to update.</div>
    <input class="search" id="procSearch" placeholder="Filter proceedings…">
    <div class="grid" id="procGrid"></div>
  </section>

  <section id="digest">
    <div class="note"><b>Tracked developments</b> from the spreadsheet's Digest Log. The weekday digest drafts leads; confirmed items land here. R.25-04-010 rows are highlighted.</div>
    <input class="search" id="digSearch" placeholder="Filter developments…">
    <table><thead><tr><th>Date</th><th>Source / Docket</th><th>Type</th><th>Headline</th><th>Impact</th><th>Relevance</th><th>Action</th><th>Link</th></tr></thead><tbody id="digBody"></tbody></table>
    <div class="empty" id="digEmpty" style="display:none">No developments logged yet. They'll appear here as you fill the Digest Log tab.</div>
  </section>

  <section id="agency">
    <div class="note"><b>Non-CPUC sources.</b> Check the page, then log relevant items to the Digest Log. The subscription behind each source is the safety net.</div>
    <div class="grid" id="agencyGrid"></div>
  </section>

  <section id="subs">
    <div class="note"><b>Reliability backbone.</b> The CPUC docket system is JavaScript-rendered and can't be reliably auto-scraped — the official subscriptions can, and never miss a filing. Set these up once.</div>
    <table><thead><tr><th>Source</th><th>Covers</th><th>Subscribe</th><th>Status</th><th>Notes</th></tr></thead><tbody id="subsBody"></tbody></table>
  </section>

  <footer>Verify CONFIRM # dockets with the Director. Docket links open official state systems.</footer>
</div>
<script>
const DATA = /*__DATA__*/ {};
document.getElementById('upd').textContent = DATA.updated || '';
function depthClass(d){d=(d||'').toUpperCase();if(d.includes('FULL'))return['full','FULL READ'];if(d.includes('PRIORITY'))return['prio','SCAN · PRIORITY'];if(d.includes('CONFIRM'))return['confirm','CONFIRM #'];return['scan','SCAN']}
function esc(s){return (s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;')}
function renderProc(filter){
  const g=document.getElementById('procGrid');g.innerHTML='';
  DATA.proceedings.filter(p=>!filter||JSON.stringify(p).toLowerCase().includes(filter)).forEach(p=>{
    const[cls,lab]=depthClass(p.depth);
    g.innerHTML+=`<div class="card"><div class="docket">${esc(p.docket)}</div><h3>${esc(p.title)}</h3>
      <div><span class="tag ${cls}">${lab}</span></div><div class="cat">${esc(p.category)} · ${esc(p.status)}</div>
      ${p.qual?`<div class="desc">${esc(p.qual)}</div>`:''}
      ${p.dates?`<div class="kv"><b>Upcoming:</b> ${esc(p.dates)}</div>`:''}
      ${p.impact?`<div class="kv"><b>Impact:</b> ${esc(p.impact)}</div>`:''}
      ${p.relevance?`<div class="kv"><b>Relevance:</b> ${esc(p.relevance)}</div>`:''}
      ${p.url?`<a class="open" href="${p.url}" target="_blank">Open live docket &rarr;</a>`:`<div class="sub">Confirm docket number first.</div>`}
    </div>`;});
}
function renderDigest(filter){
  const b=document.getElementById('digBody');b.innerHTML='';
  let items=DATA.digest.filter(d=>d.headline&&(!filter||JSON.stringify(d).toLowerCase().includes(filter)));
  items.sort((a,z)=>(z.date||'').localeCompare(a.date||''));
  document.getElementById('digEmpty').style.display=items.length?'none':'block';
  items.forEach(d=>{
    const prio=(d.docket||'').includes('25-04-010')||(d.source||'').includes('25-04-010');
    const rel=(d.relevance||'').trim();
    b.innerHTML+=`<tr class="${prio?'prio-row':''}"><td>${esc(d.date)}</td><td>${esc(d.source)}${d.docket&&d.docket!==d.source?'<br><small>'+esc(d.docket)+'</small>':''}</td>
      <td>${esc(d.type)}</td><td>${esc(d.headline)}</td><td>${esc(d.impact)}</td>
      <td>${rel?`<span class="pill ${rel}">${esc(rel)}</span>`:''}</td><td>${esc(d.action)}</td>
      <td>${d.link&&d.link.startsWith('http')?`<a href="${d.link}" target="_blank">open</a>`:''}</td></tr>`;});
}
function renderAgency(){
  const g=document.getElementById('agencyGrid');g.innerHTML='';
  DATA.watch.forEach(a=>{
    const isUrl=(a.page||'').startsWith('http');
    const subUrl=(a.sub||'').match(/https?:\/\/\S+/);
    g.innerHTML+=`<div class="card"><h3>${esc(a.source)}</h3><div class="desc">${esc(a.monitor)}</div>
      ${a.latest?`<div class="kv"><b>Latest:</b> ${esc(a.latest)}</div>`:''}
      ${isUrl?`<a class="open" href="${a.page}" target="_blank">Check page &rarr;</a>`:''}
      <div class="sub">Subscribe: ${subUrl?`<a href="${subUrl[0]}" target="_blank">${esc(a.sub)}</a>`:esc(a.sub)}</div></div>`;});
}
function renderSubs(){
  const b=document.getElementById('subsBody');b.innerHTML='';
  DATA.subs.forEach(s=>{const u=(s.url||'').startsWith('http');
    b.innerHTML+=`<tr><td><b>${esc(s.source)}</b></td><td>${esc(s.covers)}</td>
      <td>${u?`<a href="${s.url}" target="_blank">Subscribe</a>`:esc(s.url)}</td><td>${esc(s.status)}</td><td>${esc(s.notes)}</td></tr>`;});
}
renderProc('');renderDigest('');renderAgency();renderSubs();
document.getElementById('procSearch').oninput=e=>renderProc(e.target.value.toLowerCase());
document.getElementById('digSearch').oninput=e=>renderDigest(e.target.value.toLowerCase());
document.querySelectorAll('.tab').forEach(t=>t.onclick=()=>{
  document.querySelectorAll('.tab').forEach(x=>x.classList.remove('active'));
  document.querySelectorAll('section').forEach(x=>x.classList.remove('active'));
  t.classList.add('active');document.getElementById(t.dataset.t).classList.add('active');});
</script>
</body>
</html>"""

def apex(num):
    if not num or not num.strip().upper().startswith("R"):
        return None
    core = num.replace("R.", "R").replace(".", "").replace("-", "").strip()
    return f"https://apps.cpuc.ca.gov/apex/f?p=401:56:0::NO:RP,57,RIR:P5_PROCEEDING_SELECT:{core}"

def rows(ws, header_row, ncols):
    out = []
    r = header_row + 1
    while r <= ws.max_row:
        vals = [ws.cell(row=r, column=c).value for c in range(1, ncols + 1)]
        if all(v is None or str(v).strip() == "" for v in vals):
            r += 1
            if r > header_row + 60:
                break
            continue
        out.append([("" if v is None else str(v).strip()) for v in vals])
        r += 1
    return out

wb = load_workbook(XLSX, data_only=True)

# Proceedings
p = wb["CPUC Proceedings"]
proceedings = []
for v in rows(p, 4, 12):
    docket = v[0]
    proceedings.append({
        "docket": docket, "title": v[1], "category": v[2], "depth": v[3],
        "status": v[4], "dates": v[5], "checked": v[6], "qual": v[7],
        "quant": v[8], "relevance": v[9], "impact": v[10],
        "url": apex(docket) if docket and docket != "CONFIRM #" else None,
    })

# Watchlist
w = wb["Agency Watchlist"]
watch = []
for v in rows(w, 4, 7):
    watch.append({"source": v[0], "monitor": v[1], "page": v[2], "sub": v[3],
                  "checked": v[4], "latest": v[5], "note": v[6]})

# Digest Log
d = wb["Digest Log"]
digest = []
for v in rows(d, 4, 11):
    digest.append({"date": v[0], "source": v[1], "docket": v[2], "type": v[3],
                   "headline": v[4], "impact": v[5], "relevance": v[6],
                   "action": v[7], "owner": v[8], "reported": v[9], "link": v[10]})

# Subscriptions
s = wb["Subscriptions"]
subs = []
for v in rows(s, 4, 5):
    subs.append({"source": v[0], "covers": v[1], "url": v[2], "status": v[3], "notes": v[4]})

data = {"proceedings": proceedings, "watch": watch, "digest": digest,
        "subs": subs, "updated": datetime.date.today().isoformat()}

html = TEMPLATE.replace("/*__DATA__*/ {}", json.dumps(data, ensure_ascii=False))
with open(OUT, "w", encoding="utf-8") as f:
    f.write(html)
print("Wrote", OUT, "-", len(proceedings), "proceedings,", len(digest), "digest rows")
